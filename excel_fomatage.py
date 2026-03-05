






from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type="solid")
header_align = Alignment(horizontal='center', vertical='center')


thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

def excel_formatage(file, target_col='efficiency', seuil_rouge=0.12, seuil_vert=0.05):

    """
        Met en forme automatiquement un fichier Excel multi-onglets :
      - En-têtes stylées (couleur, gras, centrées)
      - Colonnes auto-ajustées
      - Formatage conditionnel sur une colonne cible
      - Bordures fines pour une meilleure lisibilité
    """


    wb = load_workbook(file)
    #ws = wb.active

    # Parcourir toutes les feuilles du classeur
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"#figer l'entete 

        print(f"📊 Ajustement des colonnes pour : {sheet_name}")
        # --- Mise en forme de l'en-tête ---
        for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
        
        # Ajuster chaque colonne de cette feuille
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            # Trouver la longueur max dans la colonne
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

                # Appliquer la largeur ajustée
            adjusted_width = min(max_length + 2, 50)  # Maximum 50 de large
            ws.column_dimensions[column_letter].width = adjusted_width

        # --- Recherche de la colonne cible ---
        col_letter = None      
        for cell in ws[1]:
            if str(cell.value).lower() == target_col:
                col_letter = get_column_letter(cell.column)
                col_index = cell.column
                break

        if not col_letter:
            print(f"⚠️ Colonne '{target_col}' introuvable dans {sheet_name}")
            continue

        #On applique une bordure sur les cellules colores pour garder une meilleur visibilité
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=cell.column, max_col=cell.column):
            for cell in row:
                cell.border = thin_border

        # --- Formatage conditionnel (après les bordures) ---
        data_range = f"{col_letter}2:{col_letter}{ws.max_row}"

        red_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
        rule_red = CellIsRule(operator='greaterThan', formula=[seuil_rouge], fill=red_fill)

        green_fill = PatternFill(start_color='D4EFDF', end_color='D4EFDF', fill_type='solid')
        rule_green = CellIsRule(operator='lessThanOrEqual', formula=[seuil_vert], fill=green_fill)
            # Appliquer les deux règles à la colonne Calories
        ws.conditional_formatting.add(data_range, rule_red)
        ws.conditional_formatting.add(data_range, rule_green)        



            
    wb.save(file)
    print(f"{len(wb.sheetnames)} formaté avec succées")
    print(f"✅ Fichier Excel multi-onglets créé : \n{file}")